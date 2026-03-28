"""
B6x Excel Parser - Specialized parsers for B6x SDK Excel format
================================================================

This module contains specialized parsers for the actual B6x Excel file formats:
- Flash-Map.xlsx: Documentary format with descriptive memory information
- SRAM allocation.xlsx: Complex sparse table with multiple library types

Author: B6x MCP Server Team
Version: 1.0.0
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass

logger = logging.getLogger(__name__)


class B6xExcelParser:
    """Specialized parser for B6x SDK Excel files."""

    def __init__(self, openpyxl=None, pandas=None):
        """Initialize parser with optional library backends."""
        self.openpyxl = openpyxl
        self.pd = pandas
        self._openpyxl_available = openpyxl is not None
        self._pandas_available = pandas is not None

    # ============================================================================
    # Flash Map Parser (Documentary Format)
    # ============================================================================

    def parse_flash_map_b6x_format(self, xlsx_path: str) -> List[dict]:
        """
        Parse Flash-Map.xlsx in B6x documentary format.

        The B6x Flash-Map.xlsx uses a documentary format with:
        - Descriptive text headers (Size: 2Mb(256KB), Range: 000000H - 03FFFFH)
        - Multiple sheets: "Flash 256KB", "Flash 128KB", "DATA_FLASH"

        Returns:
            List of memory region dictionaries
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"Flash map file not found: {xlsx_path}")
            return []

        regions = []

        try:
            if not self._openpyxl_available:
                logger.error("openpyxl is required for B6x Flash map parsing")
                return []

            wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            # Parse each relevant sheet
            for sheet_name in wb.sheetnames:
                if 'Flash' in sheet_name or 'DATA' in sheet_name:
                    sheet_regions = self._parse_flash_sheet_b6x(wb[sheet_name], sheet_name)
                    regions.extend(sheet_regions)

            wb.close()

            logger.info(f"Parsed {len(regions)} memory regions from B6x Flash map format")
            return regions

        except Exception as e:
            logger.error(f"Failed to parse B6x flash map {xlsx_path}: {e}")
            return []

    def _parse_flash_sheet_b6x(self, sheet, sheet_name: str) -> List[dict]:
        """Parse a single Flash sheet in B6x documentary format."""
        regions = []
        rows = list(sheet.iter_rows(values_only=True))

        # Extract information from documentary format
        current_region = None

        for row in rows:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Look for key patterns in the first column
            first_cell = str(row[0]).strip() if row[0] else ""

            # Pattern: "Size: 2Mb(256KB)" or "Size: 1Mb(128KB)"
            if first_cell.startswith("Size:"):
                size_match = re.search(r'(\d+)KB', first_cell)
                if size_match and not current_region:
                    size_kb = int(size_match.group(1))
                    current_region = {
                        "region_name": sheet_name.replace(" ", "_"),
                        "size": size_kb * 1024,
                        "size_kb": size_kb
                    }

            # Pattern: "Range: 000000H - 03FFFFH"
            elif first_cell.startswith("Range:"):
                range_match = re.search(r'([0-9A-Fa-f]+)H\s*-\s*([0-9A-Fa-f]+)H', first_cell)
                if range_match and current_region:
                    start_addr = int(range_match.group(1), 16)
                    end_addr = int(range_match.group(2), 16)
                    current_region["base_address"] = f"0x{range_match.group(1).upper()}"
                    current_region["end_address"] = f"0x{range_match.group(2).upper()}"

                    # Determine if XIP based on address
                    current_region["is_xip"] = start_addr < 0x10000000  # Flash region
                    current_region["access_type"] = "READ_ONLY" if current_region["is_xip"] else "READ_WRITE"
                    current_region["description"] = f"{sheet_name} Flash memory"

            # Pattern: "Page: 256Bytes", "Sector: 4KB", "Block: 32/64KB"
            elif first_cell.startswith("Page:") or first_cell.startswith("Sector:") or first_cell.startswith("Block:"):
                if current_region:
                    # Add geometry info to description
                    if "geometry" not in current_region:
                        current_region["geometry"] = []
                    current_region["geometry"].append(first_cell)

            # Finalize region when we have enough information
            if current_region and "base_address" in current_region:
                regions.append(current_region)
                current_region = None

        # Fallback: if no regions parsed from documentary format, create from sheet name
        if not regions:
            if "256KB" in sheet_name:
                regions.append({
                    "region_name": "FLASH",
                    "base_address": "0x08000000",
                    "size": 262144,
                    "size_kb": 256,
                    "is_xip": True,
                    "access_type": "READ_ONLY",
                    "description": "Code Flash (XIP supported)"
                })
            elif "128KB" in sheet_name:
                regions.append({
                    "region_name": "FLASH_128KB",
                    "base_address": "0x08000000",
                    "size": 131072,
                    "size_kb": 128,
                    "is_xip": True,
                    "access_type": "READ_ONLY",
                    "description": "Code Flash 128KB variant"
                })
            elif "DATA" in sheet_name.upper():
                regions.append({
                    "region_name": "DATA_FLASH",
                    "base_address": "0x18000000",
                    "size": 16384,
                    "size_kb": 16,
                    "is_xip": False,
                    "access_type": "READ_WRITE",
                    "description": "Data Flash for configuration and OTA"
                })

        return regions

    # ============================================================================
    # SRAM Allocation Parser (Sparse Table Format)
    # ============================================================================

    def parse_sram_allocation_b6x_format(self, xlsx_path: str) -> Tuple[List[dict], List[dict]]:
        """
        Parse SRAM allocation in B6x sparse table format.

        The B6x SRAM allocation Excel has:
        - Header row with library types (主从共3个连接, 单从, etc.)
        - Sparse table with addresses and sizes for different regions
        - Multiple SRAM regions (SRAM1, SRAM2, SRAM3)

        Returns:
            Tuple of (SRAM regions list, memory boundaries list)
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"SRAM allocation file not found: {xlsx_path}")
            return [], []

        regions = []
        boundaries = []

        try:
            if not self._openpyxl_available:
                logger.error("openpyxl is required for B6x SRAM parsing")
                return [], []

            wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            # Find the SRAM空间分配 sheet
            sheet = None
            for sheet_name in wb.sheetnames:
                if 'SRAM' in sheet_name or 'sram' in sheet_name:
                    sheet = wb[sheet_name]
                    break

            if not sheet:
                sheet = wb.active

            regions, boundaries = self._parse_sram_sheet_b6x(sheet)

            wb.close()

            logger.info(f"Parsed {len(regions)} SRAM regions and {len(boundaries)} boundaries from B6x format")
            return regions, boundaries

        except Exception as e:
            logger.error(f"Failed to parse B6x SRAM allocation {xlsx_path}: {e}")
            return [], []

    def _parse_sram_sheet_b6x(self, sheet) -> Tuple[List[dict], List[dict]]:
        """Parse SRAM sheet in B6x sparse table format."""
        regions = []
        boundaries = []

        rows = list(sheet.iter_rows(values_only=True))

        # Find library types from header row (row 0)
        library_types = []
        if rows:
            header_row = rows[0]
            # Extract library types from columns 4, 7, 10, 13...
            for i in range(3, len(header_row), 3):
                lib_type = str(header_row[i]).strip() if header_row[i] else ""
                if lib_type and '连接' in lib_type:
                    # Map to standard library names
                    if '3个' in lib_type:
                        library_types.append('ble6')
                    elif '1个' in lib_type and 'lite' in lib_type.lower():
                        library_types.append('ble6_lite')
                    elif '6个' in lib_type:
                        library_types.append('ble6_8act_6con')
                    else:
                        library_types.append(lib_type)

        logger.debug(f"Found library types: {library_types}")

        # Parse SRAM regions
        current_sram = None
        current_address = None
        region_data = {lib: [] for lib in library_types}

        for row_idx, row in enumerate(rows):
            if not row or all(cell is None or (str(cell).strip() == '' if cell else True) for cell in row):
                continue

            # Check for SRAM region marker (SRAM1, SRAM2, SRAM3)
            first_cell = str(row[0]).strip() if row[0] else ""
            if first_cell.startswith('SRAM'):
                current_sram = first_cell
                continue

            # Check for address marker (column 2)
            if len(row) > 2 and row[2]:
                addr_str = str(row[2]).strip()
                if addr_str.startswith('0x') or addr_str.startswith('0X'):
                    try:
                        current_address = int(addr_str, 16)
                    except ValueError:
                        pass

            # Extract region info for each library type
            for lib_idx, lib_type in enumerate(library_types):
                col_idx = 4 + (lib_idx * 3)
                if col_idx < len(row) and row[col_idx]:
                    cell_value = str(row[col_idx]).strip()

                    # Extract size information (e.g., "0x1D68", "用户自定义\n0x1D68")
                    size_match = re.search(r'0x([0-9A-Fa-f]+)', cell_value)
                    if size_match and current_address:
                        size = int(size_match.group(1), 16)

                        # Extract region name from cell value
                        region_name = cell_value.split('\n')[0].strip()
                        if not region_name or region_name.startswith('0x'):
                            region_name = f"{current_sram}_REGION_{len(region_data[lib_type])}"

                        region_data[lib_type].append({
                            'name': region_name,
                            'sram': current_sram,
                            'start_address': current_address,
                            'size': size,
                            'library': lib_type
                        })

        # Convert parsed data to standard format
        for lib_type in library_types:
            lib_regions = region_data.get(lib_type, [])
            for region in lib_regions:
                start_addr = region['start_address']
                size = region['size']

                regions.append({
                    'region_name': region['name'],
                    'start_address': f"0x{start_addr:X}",
                    'end_address': f"0x{start_addr + size:X}",
                    'size_bytes': size,
                    'size_kb': size // 1024,
                    'description': f"{region['name']} ({lib_type})",
                    'library_type': lib_type
                })

                # Add boundary information
                boundaries.append({
                    'region_name': region['name'],
                    'start_address': f"0x{start_addr:X}",
                    'end_address': f"0x{start_addr + size:X}",
                    'size_bytes': size,
                    'size_kb': size // 1024,
                    'boundary_type': 'allocated',
                    'library_type': lib_type,
                    'description': f"{region['name']} boundary for {lib_type}",
                    'sram_region': region['sram']
                })

        return regions, boundaries

    # ============================================================================
    # Power Consumption Parser
    # ============================================================================

    def parse_power_consumption_b6x_format(self, xlsx_path: str) -> List[dict]:
        """
        Parse power consumption data from B6x format.

        B6x Excel format:
        - Column 0: Mode (睡眠, 广播, 连接, 扫描) - may be merged cells
        - Column 1: Description (Sleep模式, 快速广播, etc.)
        - Column 2: 3V 电流 (70uA, 3.3mA, etc.)
        - Column 5: 低功耗(uA) - calculated average
        - Column 6: 峰值电流(mA)
        - Column 7: 备注

        Returns:
            List of power consumption entries
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"Power consumption file not found: {xlsx_path}")
            return []

        entries = []

        try:
            if not self._openpyxl_available:
                logger.error("openpyxl is required for B6x power parsing")
                return []

            wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            # Find the sheet with power consumption data
            # Try sheets with "功耗" or "B6x" in name, or use first sheet
            sheet = None
            for sheet_name in wb.sheetnames:
                # Skip "进入低功耗的方式" sheet which has no data
                if '方式' in sheet_name or 'method' in sheet_name.lower():
                    continue
                if '功耗' in sheet_name or 'B6x' in sheet_name:
                    sheet = wb[sheet_name]
                    break

            if sheet is None:
                sheet = wb.worksheets[0]  # Use first sheet as fallback

            rows = list(sheet.iter_rows(values_only=True))

            # Find header row - look for "3V" or "电流" or "功耗" in any cell
            header_row_idx = None
            for idx, row in enumerate(rows):
                if row and any(cell and ('3V' in str(cell) or '电流' in str(cell) or '低功耗' in str(cell)) for cell in row):
                    header_row_idx = idx
                    break

            if header_row_idx is None:
                logger.warning("Could not find power consumption header row")
                wb.close()
                return []

            # Parse data rows - column indices based on actual Excel structure
            # Col 0: Mode, Col 1: Description, Col 2: 3V Current, Col 5: Avg uA, Col 6: Peak mA, Col 7: Notes
            for row in rows[header_row_idx + 1:]:
                if not row or len(row) < 3:
                    continue

                # Get mode from col 0 (may be empty due to merged cells)
                mode = str(row[0]).strip() if row[0] and str(row[0]).strip().lower() not in ['nan', 'none', ''] else ""
                description = str(row[1]).strip() if len(row) > 1 and row[1] and str(row[1]).strip().lower() not in ['nan', 'none'] else ""
                current_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                avg_ua = row[5] if len(row) > 5 and row[5] else None
                peak_ma = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                notes = str(row[7]).strip() if len(row) > 7 and row[7] and str(row[7]).strip().lower() not in ['nan', 'none'] else ""

                # Skip empty rows
                if not mode and not description and not current_str:
                    continue

                # Build mode from description if col 0 is empty (merged cells)
                if not mode and description:
                    # Infer mode from description
                    desc_lower = description.lower()
                    if 'sleep' in desc_lower or 'poweroff' in desc_lower:
                        mode = "睡眠"
                    elif '广播' in description or 'adv' in desc_lower:
                        mode = "广播"
                    elif '连接' in description or 'conn' in desc_lower:
                        mode = "连接"
                    elif '扫描' in description or 'scan' in desc_lower:
                        mode = "扫描"
                    else:
                        mode = "其他"

                # Parse current value from col 2 (e.g., "70uA", "3.3mA")
                current_ua = 0
                if current_str and current_str.lower() not in ['nan', 'none', '']:
                    current_match = re.search(r'([\d.]+)\s*([mu]?)A', current_str, re.IGNORECASE)
                    if current_match:
                        current_val = float(current_match.group(1))
                        unit = current_match.group(2).lower()
                        # Convert to microamps
                        if unit == 'm':
                            current_ua = current_val * 1000
                        else:
                            current_ua = current_val

                # Use avg_ua from col 5 if available and current_ua is 0
                if current_ua == 0 and avg_ua is not None:
                    try:
                        current_ua = float(avg_ua)
                    except (ValueError, TypeError):
                        pass

                # Skip if no valid data
                if not mode or current_ua == 0:
                    continue

                entries.append({
                    'mode': mode,
                    'description': description,
                    'current_ua': current_ua,
                    'voltage_mv': 3300,  # Default 3.3V
                    'power_mw': (current_ua / 1000) * 3.3,
                    'peripheral': 'SYSTEM',
                    'peak_current_ma': peak_ma,
                    'conditions': notes
                })

            wb.close()

            logger.info(f"Parsed {len(entries)} power consumption entries")
            return entries

        except Exception as e:
            logger.error(f"Failed to parse B6x power consumption {xlsx_path}: {e}")
            return []


def create_b6x_parser(openpyxl=None, pandas=None):
    """Factory function to create B6x Excel parser."""
    return B6xExcelParser(openpyxl=openpyxl, pandas=pandas)


# Standalone test functions
def parse_flash_map_b6x(xlsx_path: str) -> List[dict]:
    """Standalone function to parse B6x Flash map."""
    try:
        import openpyxl
        parser = B6xExcelParser(openpyxl=openpyxl)
        return parser.parse_flash_map_b6x_format(xlsx_path)
    except ImportError:
        logger.error("openpyxl is required")
        return []


def parse_sram_allocation_b6x(xlsx_path: str) -> Tuple[List[dict], List[dict]]:
    """Standalone function to parse B6x SRAM allocation."""
    try:
        import openpyxl
        parser = B6xExcelParser(openpyxl=openpyxl)
        return parser.parse_sram_allocation_b6x_format(xlsx_path)
    except ImportError:
        logger.error("openpyxl is required")
        return [], []


def parse_power_consumption_b6x(xlsx_path: str) -> List[dict]:
    """Standalone function to parse B6x power consumption."""
    try:
        import openpyxl
        parser = B6xExcelParser(openpyxl=openpyxl)
        return parser.parse_power_consumption_b6x_format(xlsx_path)
    except ImportError:
        logger.error("openpyxl is required")
        return []
