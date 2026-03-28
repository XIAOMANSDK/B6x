"""
Excel to Markdown Converter - Stage 1 of Excel Parsing Pipeline
================================================================

Convert Excel files to Markdown format for easier table parsing.

This module provides:
- MarkItDown-based Excel to Markdown conversion (primary)
- openpyxl-based conversion (fallback)
- Caching support for incremental builds
- Chinese character preservation

Usage:
    converter = ExcelToMarkdownConverter(cache_dir)
    md_path = converter.convert_excel_to_md(excel_path)

Author: B6x MCP Server Team
Version: 1.0.0
"""

import os
import logging
import hashlib
from pathlib import Path
from typing import Optional, Dict, Any
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelToMarkdownConverter:
    """
    Convert Excel files to Markdown with caching support.

    Primary method: MarkItDown (handles complex formats, Chinese characters)
    Fallback method: openpyxl (basic table conversion)
    """

    # Files that require Markdown pipeline (known problematic formats)
    MARKDOWN_PIPELINE_FILES = [
        "B6x-Flash-Map.xlsx",
        "B6x_BLE-Sram空间分配.xlsx",
        "B6x BLE兼容性列表.xlsx"
    ]

    def __init__(self, cache_dir: Path, cache_manager=None):
        """
        Initialize the converter.

        Args:
            cache_dir: Directory for Markdown cache files
            cache_manager: Optional cache manager for incremental builds
        """
        self.cache_dir = Path(cache_dir)
        self.cache_manager = cache_manager
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        # Initialize MarkItDown
        self._markitdown_available = False
        self._openpyxl_available = False

        try:
            from markitdown import MarkItDown
            self.MarkItDown = MarkItDown
            self._markitdown_available = True
            logger.info("MarkItDown available for Excel conversion")
        except ImportError:
            logger.warning("MarkItDown not available for Excel conversion")

        try:
            import openpyxl
            self.openpyxl = openpyxl
            self._openpyxl_available = True
            logger.info("openpyxl available as fallback")
        except ImportError:
            logger.warning("openpyxl not available as fallback")

    def should_use_markdown_pipeline(self, excel_filename: str) -> bool:
        """
        Check if file should use Markdown pipeline.

        Args:
            excel_filename: Name of Excel file

        Returns:
            True if file is in the known problematic list
        """
        return excel_filename in self.MARKDOWN_PIPELINE_FILES

    def convert_excel_to_md(
        self,
        excel_path: Path,
        force: bool = False
    ) -> Optional[Path]:
        """
        Convert Excel file to Markdown with caching support.

        Args:
            excel_path: Path to Excel file
            force: Force re-conversion even if cached

        Returns:
            Path to Markdown file, or None if conversion failed
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            return None

        # Generate cache filename
        md_filename = excel_path.stem + ".md"
        md_path = self.cache_dir / md_filename

        # Check cache if not forcing rebuild
        if not force and md_path.exists():
            if self._is_cache_valid(excel_path, md_path):
                logger.debug(f"Using cached Markdown: {md_path}")
                return md_path
            else:
                logger.debug(f"Cache invalid, re-converting: {excel_path}")

        # Perform conversion
        logger.info(f"Converting {excel_path.name} to Markdown...")

        # Try MarkItDown first
        if self._markitdown_available:
            md_path = self._convert_with_markitdown(excel_path)
            if md_path:
                self._save_cache_metadata(excel_path, md_path)
                return md_path

        # Fallback to openpyxl
        if self._openpyxl_available:
            md_path = self._convert_with_openpyxl(excel_path)
            if md_path:
                self._save_cache_metadata(excel_path, md_path)
                return md_path

        logger.error(f"Failed to convert {excel_path} to Markdown")
        return None

    def _convert_with_markitdown(self, excel_path: Path) -> Optional[Path]:
        """
        Convert Excel to Markdown using MarkItDown (primary method).

        MarkItDown handles:
        - Complex table structures
        - Chinese characters correctly
        - Sparse tables
        - Merged cells

        Args:
            excel_path: Path to Excel file

        Returns:
            Path to Markdown file, or None if failed
        """
        try:
            md = self.MarkItDown()
            result = md.convert(str(excel_path))

            # Save to cache
            md_filename = excel_path.stem + ".md"
            md_path = self.cache_dir / md_filename

            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(result.text_content)

            logger.info(f"MarkItDown conversion successful: {md_path}")
            return md_path

        except Exception as e:
            logger.error(f"MarkItDown conversion failed for {excel_path}: {e}")
            return None

    def _convert_with_openpyxl(self, excel_path: Path) -> Optional[Path]:
        """
        Convert Excel to Markdown using openpyxl (fallback method).

        This is a simpler conversion that may not handle all edge cases.

        Args:
            excel_path: Path to Excel file

        Returns:
            Path to Markdown file, or None if failed
        """
        try:
            wb = self.openpyxl.load_workbook(
                excel_path,
                read_only=True,
                data_only=True
            )

            md_lines = [f"# {excel_path.stem}\n"]

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                md_lines.append(f"\n## Sheet: {sheet_name}\n")

                # Convert to Markdown table
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    # Skip completely empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue

                    # Convert row to Markdown table row
                    md_row = "| " + " | ".join(
                        str(cell) if cell is not None else ""
                        for cell in row
                    ) + " |"
                    md_lines.append(md_row)

                    # Add separator after first non-empty row (header)
                    if row_idx == 0 or (row_idx == 1 and len(md_lines) == 3):
                        separator = "| " + " | ".join(
                            "---" for _ in row
                        ) + " |"
                        md_lines.append(separator)

            wb.close()

            # Save to cache
            md_filename = excel_path.stem + ".md"
            md_path = self.cache_dir / md_filename

            with open(md_path, 'w', encoding='utf-8') as f:
                f.write("\n".join(md_lines))

            logger.info(f"openpyxl conversion successful: {md_path}")
            return md_path

        except Exception as e:
            logger.error(f"openpyxl conversion failed for {excel_path}: {e}")
            return None

    def _is_cache_valid(self, excel_path: Path, md_path: Path) -> bool:
        """
        Check if cached Markdown is still valid.

        Args:
            excel_path: Original Excel file
            md_path: Cached Markdown file

        Returns:
            True if cache is valid
        """
        # Check modification times
        excel_mtime = excel_path.stat().st_mtime
        md_mtime = md_path.stat().st_mtime

        return md_mtime > excel_mtime

    def _save_cache_metadata(self, excel_path: Path, md_path: Path):
        """Save cache metadata for incremental build support."""
        if self.cache_manager:
            try:
                # Calculate file hash
                with open(excel_path, 'rb') as f:
                    file_hash = hashlib.md5(f.read()).hexdigest()

                # Store metadata
                metadata = {
                    "excel_path": str(excel_path),
                    "md_path": str(md_path),
                    "file_hash": file_hash,
                    "converted_at": datetime.now().isoformat()
                }

                # Use cache manager if available
                if hasattr(self.cache_manager, 'save_cache_info'):
                    self.cache_manager.save_cache_info(str(excel_path), metadata)
            except Exception as e:
                logger.debug(f"Failed to save cache metadata: {e}")

    def get_cached_md_path(self, excel_filename: str) -> Optional[Path]:
        """
        Get path to cached Markdown file for an Excel filename.

        Args:
            excel_filename: Name of Excel file (e.g., "B6x-Flash-Map.xlsx")

        Returns:
            Path to cached Markdown file, or None if not cached
        """
        md_filename = Path(excel_filename).stem + ".md"
        md_path = self.cache_dir / md_filename

        if md_path.exists():
            return md_path
        return None

    def clear_cache(self, excel_filename: Optional[str] = None):
        """
        Clear cached Markdown files.

        Args:
            excel_filename: Specific file to clear, or None to clear all
        """
        if excel_filename:
            md_filename = Path(excel_filename).stem + ".md"
            md_path = self.cache_dir / md_filename
            if md_path.exists():
                md_path.unlink()
                logger.debug(f"Cleared cache: {md_path}")
        else:
            # Clear all cached Markdown files
            for md_file in self.cache_dir.glob("*.md"):
                md_file.unlink()
                logger.debug(f"Cleared cache: {md_file}")


# ============================================================================
# Convenience Functions
# ============================================================================

def convert_excel_to_markdown(
    excel_path: Path,
    cache_dir: Path,
    force: bool = False
) -> Optional[Path]:
    """
    Convenience function to convert Excel to Markdown.

    Args:
        excel_path: Path to Excel file
        cache_dir: Cache directory
        force: Force re-conversion

    Returns:
        Path to Markdown file, or None if failed
    """
    converter = ExcelToMarkdownConverter(cache_dir)
    return converter.convert_excel_to_md(excel_path, force=force)
