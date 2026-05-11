"""
Excel Hardware Constraint Parser
=================================

Parse Excel files in doc/SW_Spec/ to extract hardware constraints.

Extracted data:
- IO_MAP.xlsx: Pin muxing, alternate functions
- Flash-Map.xlsx: Memory layout, base addresses
- Power.xlsx: Power consumption data

Author: B6x MCP Server Team
Version: 0.1.0
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict

logger = logging.getLogger(__name__)


@dataclass
class PinMuxEntry:
    """Pin multiplexing configuration entry."""
    pin_number: str           # "P0_0", "P1_5"
    alternate_function: str   # "UART0_TX", "SPI0_SCK"
    peripheral: str           # "UART0", "SPI0"
    direction: str            # "RX", "TX", "BIDIR"
    is_default: bool          # Is this the default function on reset?
    register_config: str = "" # Register configuration if needed

    def __hash__(self):
        return hash((self.pin_number, self.alternate_function))


@dataclass
class MemoryRegion:
    """Memory region definition."""
    region_name: str          # "FLASH", "SRAM", "XIP_FLASH"
    base_address: str         # "0x08000000"
    size: int                 # Size in bytes
    size_kb: int              # Size in KB
    is_xip: bool              # Supports XIP (eXecute In Place)
    access_type: str          # "READ_WRITE", "READ_ONLY", "EXECUTE"
    description: str = ""


@dataclass
class PowerEntry:
    """Power consumption entry."""
    mode: str                 # "ACTIVE", "SLEEP", "DEEP_SLEEP"
    peripheral: str           # "UART", "GPIO", "BLE"
    current_ua: float         # Current in microamps
    voltage_mv: int           # Voltage in millivolts
    power_mw: float           # Power in milliwatts
    conditions: str = ""      # Test conditions


@dataclass
class SRAMRegion:
    """SRAM memory region for BLE stack."""
    region_name: str          # "BLE_STACK", "BLE_HEAP", "APP_HEAP"
    start_address: str        # "0x20000000"
    end_address: str          # "0x20003FFF"
    size_bytes: int           # Size in bytes
    size_kb: int              # Size in KB
    description: str = ""     # Region description
    library_type: str = ""    # "lite", "standard", "full"


@dataclass
class CompatibilityEntry:
    """BLE compatibility matrix entry."""
    feature_name: str         # "BLE 5.0", "Mesh", "OTA"
    supported_libs: List[str] # ["lite", "standard", "full"]
    min_sdk_version: str = "" # "1.0.0"
    max_sdk_version: str = "" # "" (no limit)
    notes: str = ""           # Additional notes
    constraints: List[str] = None  # Connection limits, etc.

    def __post_init__(self):
        if self.constraints is None:
            self.constraints = []


@dataclass
class PhoneCompatibilityIssue:
    """Specific phone model BLE compatibility issue (pitfall)."""
    phone_model: str          # e.g., "iPhone 12", "Huawei P40", "Xiaomi Mi 11"
    phone_brand: str = ""     # e.g., "Apple", "Huawei", "Xiaomi"
    android_version: str = "" # e.g., "Android 11", "iOS 14"
    issue_type: str = ""      # e.g., "连接失败", "音频断续", "无法配对"
    issue_description: str = ""  # Detailed problem description
    root_cause: str = ""      # e.g., "BLE 5.0兼容性", "GATT超时"
    workaround: str = ""      # Temporary solution
    solution: str = ""         # Permanent fix
    affected_ble_version: str = ""  # e.g., "BLE 5.0", "BLE 4.2"
    sdk_version_required: str = ""  # Minimum SDK version to fix
    severity: str = "medium"  # "critical", "high", "medium", "low"
    reference: str = ""       # Issue reference or source


@dataclass
class MemoryBoundary:
    """Memory boundary/limit definition for SRAM regions."""
    region_name: str          # e.g., "BLE_STACK", "BLE_HEAP", "APP_HEAP"
    start_address: str        # e.g., "0x20000000"
    end_address: str          # e.g., "0x20003FFF"
    size_bytes: int           # Size in bytes
    size_kb: int              # Size in KB
    boundary_type: str = ""   # "fixed", "configurable", "shared"
    library_type: str = ""    # "lite", "standard", "full"
    description: str = ""     # Region description
    access_restrictions: List[str] = None  # Access constraints
    overlap_warnings: List[str] = None    # Potential overlaps with other regions
    alignment_requirement: int = 4  # Alignment requirement in bytes

    def __post_init__(self):
        if self.access_restrictions is None:
            self.access_restrictions = []
        if self.overlap_warnings is None:
            self.overlap_warnings = []


class ExcelHardwareParser:
    """
    Parse Excel hardware constraint files.

    Supports:
    - IO mapping (pin multiplexing)
    - Flash/SRAM memory layout
    - Power consumption data
    """

    def __init__(self):
        """Initialize the parser."""
        self._pandas_available = False
        self._openpyxl_available = False

        # Check for pandas (preferred for complex Excel files)
        try:
            import pandas as pd
            self.pd = pd
            self._pandas_available = True
            logger.info("pandas available for Excel parsing")
        except ImportError:
            logger.warning("pandas not available, will use openpyxl directly")

        # Check for openpyxl (basic Excel support)
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self._openpyxl_available = True
            logger.info("openpyxl available for Excel parsing")
        except ImportError:
            logger.error("openpyxl not available. Install with: pip install openpyxl pandas")

    def parse_io_map(self, xlsx_path: str) -> List[PinMuxEntry]:
        """
        Parse IO_MAP.xlsx to extract pin multiplexing configurations.

        Expected Excel format:
        - Sheet: "Pin_Mux" or "IO_MAP"
        - Columns: Pin, Function, Peripheral, Direction, Default

        Args:
            xlsx_path: Path to IO_MAP.xlsx file

        Returns:
            List of PinMuxEntry objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"IO map file not found: {xlsx_path}")
            return []

        entries = []

        try:
            if self._pandas_available:
                entries = self._parse_io_map_pandas(xlsx_path)
            elif self._openpyxl_available:
                entries = self._parse_io_map_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(entries)} pin mux entries from {xlsx_path.name}")
            return entries

        except Exception as e:
            logger.error(f"Failed to parse IO map {xlsx_path}: {e}")
            return []

    def _parse_io_map_pandas(self, xlsx_path: Path) -> List[PinMuxEntry]:
        """Parse IO map using pandas (preferred method).

        B6x_IO_MAP.xlsx format:
        Row 1: Empty
        Row 2: No | Name | PAD type | default | 0 | 1 | 2 | 3 | ...
        Row 3:    |    |          |         | GPIO | CSC | special | Timer | ...
        Row 4+: 1  | PADA0 | IO RT25 | SWCLK | PA00 | CSC | SWCLK | None | ...
        """
        entries = []

        # Read the raw Excel file (skip first empty row)
 # Row 0 is empty, Row 1 is header part 1, Row 2 is header part 2
        df_raw = self.pd.read_excel(xlsx_path, header=None)

        logger.debug(f"Excel shape: {df_raw.shape}")
        logger.debug(f"First few rows:\n{df_raw.head(5)}")

        # Find the header rows
        header_row1_idx = None
        header_row2_idx = None

        for idx, row in df_raw.iterrows():
            row_str = str(row.to_list())
            # First header row contains "No", "Name"
            if header_row1_idx is None and any('No' in str(cell) for cell in row):
                header_row1_idx = idx
            # Second header row contains "GPIO", "CSC"
            elif header_row1_idx is not None and header_row2_idx is None and any('GPIO' in str(cell) for cell in row):
                header_row2_idx = idx
                break

        if header_row1_idx is None or header_row2_idx is None:
            logger.warning(f"Could not find header rows in {xlsx_path.name}")
            return []

        logger.debug(f"Header rows at: {header_row1_idx}, {header_row2_idx}")

        # Build column mapping from the two header rows
        header1 = df_raw.iloc[header_row1_idx].fillna('').astype(str).str.strip()
        header2 = df_raw.iloc[header_row2_idx].fillna('').astype(str).str.strip()

        # Find key columns
        col_name = None
        col_default = None
        func_cols = []  # List of (col_idx, func_type) tuples

        for idx, (h1, h2) in enumerate(zip(header1, header2)):
            # Name column
            if h1 == 'Name' or h2 == 'Name':
                col_name = idx
            # Default column
            elif 'defualt' in h1.lower() or 'default' in h1.lower():
                col_default = idx
            # Function columns (numeric in header1, function type in header2)
            elif h1.isdigit() and h2 and h2 not in ['', 'None', 'nan']:
                func_cols.append((idx, h2))

        if col_name is None:
            logger.warning("Could not find 'Name' column")
            return []

        logger.debug(f"Name column: {col_name}, Default column: {col_default}")
        logger.debug(f"Function columns: {func_cols[:5]}... (total {len(func_cols)})")

        # Parse data rows (starting after header_row2)
        for idx in range(header_row2_idx + 1, len(df_raw)):
            row = df_raw.iloc[idx]

            # Get pin name
            pin_name = str(row[col_name]).strip() if col_name is not None and self.pd.notna(row[col_name]) else ""
            if not pin_name or pin_name.lower() in ['nan', '', 'none']:
                continue

            # Get default value
            default_val = str(row[col_default]).strip() if col_default is not None and self.pd.notna(row[col_default]) else ""
            is_default = False

            # Extract functions from all function columns
            for col_idx, func_type in func_cols:
                func_value = str(row[col_idx]).strip() if self.pd.notna(row[col_idx]) else ""

                # Skip empty values
                if not func_value or func_value.lower() in ['nan', '', 'none', 'n/a']:
                    continue

                # Determine if this is the default function
                if default_val and func_value in default_val:
                    is_default = True

                # Create pin mux entry
                entry = PinMuxEntry(
                    pin_number=pin_name,
                    alternate_function=func_value,
                    peripheral=self._extract_peripheral_from_function(func_value),
                    direction=self._infer_direction_from_function(func_value, func_type),
                    is_default=is_default
                )
                entries.append(entry)

        logger.info(f"Parsed {len(entries)} pin mux entries from {len(df_raw) - header_row2_idx - 1} data rows")
        return entries

    def _parse_io_map_openpyxl(self, xlsx_path: Path) -> List[PinMuxEntry]:
        """Parse IO map using openpyxl (fallback method).

        B6x_IO_MAP.xlsx format:
        Row 1: Empty
        Row 2: No | Name | PAD type | default | 0 | 1 | 2 | 3 | ...
        Row 3:    |    |          |         | GPIO | CSC | special | Timer | ...
        Row 4+: 1  | PADA0 | IO RT25 | SWCLK | PA00 | CSC | SWCLK | None | ...
        """
        entries = []

        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))

        # Find header rows
        header_row1_idx = None
        header_row2_idx = None

        for idx, row in enumerate(rows):
            row_str = [str(cell) if cell else '' for cell in row]
            # First header row contains "No", "Name"
            if header_row1_idx is None and any('No' in cell for cell in row_str):
                header_row1_idx = idx
            # Second header row contains "GPIO", "CSC"
            elif header_row1_idx is not None and header_row2_idx is None and any('GPIO' in cell for cell in row_str):
                header_row2_idx = idx
                break

        if header_row1_idx is None or header_row2_idx is None:
            wb.close()
            logger.warning("Could not find header rows in IO map")
            return []

        # Build column mapping
        header1 = [str(cell) if cell else '' for cell in rows[header_row1_idx]]
        header2 = [str(cell) if cell else '' for cell in rows[header_row2_idx]]

        col_name = None
        col_default = None
        func_cols = []

        for idx, (h1, h2) in enumerate(zip(header1, header2)):
            if h1 == 'Name' or h2 == 'Name':
                col_name = idx
            elif 'defualt' in h1.lower() or 'default' in h1.lower():
                col_default = idx
            elif h1.isdigit() and h2 and h2 not in ['', 'None']:
                func_cols.append((idx, h2))

        if col_name is None:
            wb.close()
            logger.warning("Could not find 'Name' column")
            return []

        # Parse data rows
        for idx in range(header_row2_idx + 1, len(rows)):
            row = rows[idx]

            pin_name = str(row[col_name]).strip() if row[col_name] else ""
            if not pin_name or pin_name.lower() in ['nan', '', 'none']:
                continue

            default_val = str(row[col_default]).strip() if col_default is not None and row[col_default] else ""

            for col_idx, func_type in func_cols:
                func_value = str(row[col_idx]).strip() if row[col_idx] else ""

                if not func_value or func_value.lower() in ['nan', '', 'none', 'n/a']:
                    continue

                is_default = default_val and func_value in default_val

                entry = PinMuxEntry(
                    pin_number=pin_name,
                    alternate_function=func_value,
                    peripheral=self._extract_peripheral_from_function(func_value),
                    direction=self._infer_direction_from_function(func_value, func_type),
                    is_default=is_default
                )
                entries.append(entry)

        wb.close()
        logger.info(f"Parsed {len(entries)} pin mux entries using openpyxl")
        return entries

    def _extract_peripheral_from_function(self, function: str) -> str:
        """
        Extract peripheral name from alternate function name.

        Examples:
            "UART0_TX" -> "UART0"
            "SPI0_SCK" -> "SPI0"
            "GPIO_Output" -> "GPIO"
            "P0_0" -> "UNKNOWN"
            "PA00" -> "GPIO"
            "SWCLK" -> "DEBUG"
            "CSC" -> "CSC"
        """
        if not function:
            return "UNKNOWN"

        # Handle GPIO pins (PA00, PB01, etc.)
        if function.upper().startswith(('PA', 'PB', 'PC', 'PD')):
            return "GPIO"

        # Handle debug functions
        if 'SWCLK' in function.upper() or 'SWDIO' in function.upper():
            return "DEBUG"

        # Handle CSC (Clock Security/Config)
        if 'CSC' in function.upper():
            return "CSC"

        # Check for common peripheral prefixes
        for peripheral in ['UART', 'SPI', 'I2C', 'I2S', 'ADC', 'DAC', 'PWM', 'TIMER',
                          'GPIO', 'USB', 'CAN', 'SDMMC', 'QUADSPI', 'ETH', 'BLE', 'CTMR']:
            if function.upper().startswith(peripheral):
                # Get full peripheral name (e.g., "UART0" from "UART0_TX")
                idx = len(peripheral)
                while idx < len(function) and function[idx].isdigit():
                    idx += 1
                return function[:idx]

        return "UNKNOWN"

    def _infer_direction_from_function(self, function: str, func_type: str) -> str:
        """
        Infer signal direction from function name and type.

        Args:
            function: Function name (e.g., "UART0_TX", "SPI0_SCK")
            func_type: Function type from header (e.g., "GPIO", "CSC")

        Returns:
            "RX", "TX", or "BIDIR"
        """
        if not function:
            return "BIDIR"

        func_upper = function.upper()

        # TX indicators
        if any(suffix in func_upper for suffix in ['TX', 'MOSI', 'SCLK', 'SCK', 'CK', 'CLKOUT']):
            return "TX"

        # RX indicators
        if any(suffix in func_upper for suffix in ['RX', 'MISO', 'SDA']):
            return "RX"

        # Bidirectional indicators
        if any(suffix in func_upper for suffix in ['SDA', 'SCL', 'CS', 'NCS']):
            return "BIDIR"

        # Default based on function type
        if func_type == 'GPIO':
            return "BIDIR"

        return "BIDIR"

    def _find_column_mapping(self, columns, target_keywords) -> Dict[str, str]:
        """Find column names by fuzzy matching keywords.

        Returns a mapping from keyword to original column name.
        For phone/model matching, we also try to find best match.
        """
        col_map = {}
        columns_list = [str(c) for c in columns]

        for target in target_keywords:
            for col_name in columns_list:
                # Check if target keyword is contained in column name (case-insensitive)
                if target.lower() in col_name.lower():
                    col_map[target] = col_name
                    break

        # Additional heuristic: map common logical keys to found columns
        # This handles cases where 'phone' should map to '手机型号' (found via '型号')
        logical_key_mapping = {
            'phone': ['phone', 'model', '型号', '手机型号'],
            'brand': ['brand', '品牌', '手机品牌'],
            'android': ['android', '安卓', '系统', '版本'],
            'issue': ['issue', '问题', '描述'],
            'cause': ['cause', '原因'],
            'solution': ['solution', '解决'],
            'workaround': ['workaround', '临时', '方案'],
            'ble': ['ble', '蓝牙'],
            'sdk': ['sdk'],
            'severity': ['severity', '严重', '级别']
        }

        # Build enhanced mapping
        enhanced_map = {}
        for logical_key, keywords in logical_key_mapping.items():
            for kw in keywords:
                if kw in col_map:
                    enhanced_map[logical_key] = col_map[kw]
                    break
                # Direct column match
                for col_name in columns_list:
                    if kw.lower() in col_name.lower():
                        enhanced_map[logical_key] = col_name
                        break
                if logical_key in enhanced_map:
                    break

        # Merge with original mapping (original takes precedence for non-logical keys)
        result = {**col_map, **enhanced_map}
        return result

    def parse_flash_map(self, xlsx_path: str) -> List[MemoryRegion]:
        """
        Parse Flash-Map.xlsx to extract memory layout.

        Expected Excel format:
        - Sheet: "Memory_Layout" or "Flash Map"
        - Columns: Region, Base_Address, Size_KB, XIP_Support, Access

        Args:
            xlsx_path: Path to Flash-Map.xlsx file

        Returns:
            List of MemoryRegion objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"Flash map file not found: {xlsx_path}")
            return []

        regions = []

        try:
            if self._pandas_available:
                regions = self._parse_flash_map_pandas(xlsx_path)
            elif self._openpyxl_available:
                regions = self._parse_flash_map_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(regions)} memory regions from {xlsx_path.name}")
            return regions

        except Exception as e:
            logger.error(f"Failed to parse flash map {xlsx_path}: {e}")
            return []

    def _parse_flash_map_pandas(self, xlsx_path: Path) -> List[MemoryRegion]:
        """Parse flash map using pandas."""
        regions = []

        # Try different sheet names
        for sheet_name in ['Memory_Layout', 'Flash Map', 'MemoryMap', 'Sheet1',
                           'Flash 256KB', 'Flash 128KB', 'Flash']:
            try:
                df = self.pd.read_excel(xlsx_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        else:
            df = self.pd.read_excel(xlsx_path)

        df.columns = [col.strip() for col in df.columns]

        col_map = self._find_column_mapping(df.columns,
                                            ['region', 'base', 'address', 'size', 'xip', 'access'])

        for _, row in df.iterrows():
            try:
                region_name = str(row[col_map.get('region', '')]).strip() if col_map.get('region') else ""
                base_addr = str(row[col_map.get('base', col_map.get('address', ''))]).strip() if col_map.get('base') or col_map.get('address') else ""
                size_kb = self._parse_size(row[col_map.get('size', '')]) if col_map.get('size') else 0
                is_xip = str(row[col_map.get('xip', '')]).strip().lower() in ('yes', 'true', '1') if col_map.get('xip') else False
                access = str(row[col_map.get('access', 'READ_WRITE')]).strip() if col_map.get('access') else "READ_WRITE"

                if region_name and base_addr:
                    region = MemoryRegion(
                        region_name=region_name,
                        base_address=base_addr,
                        size=size_kb * 1024,
                        size_kb=size_kb,
                        is_xip=is_xip,
                        access_type=access
                    )
                    regions.append(region)
            except Exception as e:
                logger.debug(f"Skipping row due to error: {e}")

        return regions

    def _parse_flash_map_openpyxl(self, xlsx_path: Path) -> List[MemoryRegion]:
        """Parse flash map using openpyxl with support for B6x special format."""
        regions = []

        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet_name in ['Memory_Layout', 'Flash Map', 'MemoryMap',
                           'Flash 256KB', 'Flash 128KB', 'Flash']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                break
        else:
            sheet = wb.active

        # Find header row - search in first 20 rows for B6x format
        header_row_num = None
        header_cells = None

        for row_num, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
            if row and any(cell and ('region' in str(cell).lower() or 'detail' in str(cell).lower()) for cell in row):
                header_row_num = row_num
                header_cells = row
                break

        if not header_row_num:
            # Try to find any non-empty row with structured data
            for row_num, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
                if row and len([c for c in row if c is not None]) >= 3:
                    header_row_num = row_num
                    header_cells = row
                    break

        if not header_row_num:
            wb.close()
            return []

        # Build column mapping from header
        col_map = {}
        for i, cell in enumerate(header_cells):
            if cell:
                col_lower = str(cell).lower()
                if 'region' in col_lower or 'name' in col_lower or 'detail' in col_lower:
                    col_map['region'] = i
                elif 'base' in col_lower or 'address' in col_lower:
                    col_map['base'] = i
                elif 'size' in col_lower:
                    col_map['size'] = i
                elif 'xip' in col_lower:
                    col_map['xip'] = i
                elif 'access' in col_lower:
                    col_map['access'] = i

        # Read data rows
        for row in sheet.iter_rows(min_row=header_row_num + 1, values_only=True):
            # Skip empty rows
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Extract region name (first non-empty column)
            region_name = ""
            base_addr = ""

            if col_map.get('region') is not None:
                region_name = str(row[col_map['region']]).strip() if row[col_map['region']] else ""
            else:
                # Find first non-empty cell as region name
                for cell in row:
                    if cell and str(cell).strip():
                        region_name = str(cell).strip()
                        break

            if col_map.get('base') is not None:
                base_addr = str(row[col_map['base']]).strip() if row[col_map.get('base')] else ""

            # Parse size - look for KB/MB in region description or size column
            size_kb = 0
            if col_map.get('size') is not None and row[col_map.get('size')]:
                size_kb = self._parse_size(row[col_map['size']])
            elif region_name:
                # Try to extract size from region name (e.g., "256KB")
                size_kb = self._parse_size(region_name)

            # For B6x Flash Map, extract base address from description if needed
            if not base_addr and region_name:
                # Look for address pattern like "000000H - 03FFFFH"
                import re
                addr_match = re.search(r'([0-9A-Fa-f]+)H\s*-\s*([0-9A-Fa-f]+)H', region_name)
                if addr_match:
                    base_addr = f"0x{addr_match.group(1)}"

            if region_name and (base_addr or size_kb > 0):
                region = MemoryRegion(
                    region_name=region_name,
                    base_address=base_addr or "0x08000000",
                    size=size_kb * 1024,
                    size_kb=size_kb,
                    is_xip=True,  # Flash typically supports XIP
                    access_type="READ_ONLY"
                )
                regions.append(region)

        wb.close()
        return regions

    def _parse_size(self, value) -> int:
        """Parse size value (handles KB, MB suffixes)."""
        if pd.isna(value):
            return 0

        value_str = str(value).strip().upper()

        # Extract numeric part
        num_str = ''.join(c for c in value_str if c.isdigit() or c == '.')
        try:
            num = float(num_str)
        except ValueError:
            return 0

        # Handle suffixes
        if 'KB' in value_str or 'K' in value_str:
            return int(num)
        elif 'MB' in value_str or 'M' in value_str:
            return int(num * 1024)
        else:
            # Assume KB
            return int(num)

    # ========================================================================
    # Power Consumption Parsing
    # ========================================================================

    def parse_power_map(self, xlsx_path: str) -> List[PowerEntry]:
        """
        Parse B6x_BLE-功耗参考.xlsx to extract power consumption data.

        Expected Excel format:
        - Sheet: "B6x低功耗模式" or "B6x功耗"
        - Columns: 场景, 模式, 3V平均电流, 持续时间, 间隔, 平均功耗(uA), 峰值电流, 备注

        Args:
            xlsx_path: Path to power consumption Excel file

        Returns:
            List of PowerEntry objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"Power map file not found: {xlsx_path}")
            return []

        entries = []

        try:
            if self._pandas_available:
                entries = self._parse_power_map_pandas(xlsx_path)
            elif self._openpyxl_available:
                entries = self._parse_power_map_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(entries)} power consumption entries from {xlsx_path.name}")
            return entries

        except Exception as e:
            logger.error(f"Failed to parse power map {xlsx_path}: {e}")
            return []

    def _parse_power_map_pandas(self, xlsx_path: Path) -> List[PowerEntry]:
        """Parse power map using pandas."""
        entries = []

        # Read first sheet
        df = self.pd.read_excel(xlsx_path, sheet_name=0, header=0)

        # Data structure (based on observed format):
        # Column 0: Scene (场景) - sparse
        # Column 1: Mode (模式)
        # Column 2: 3V Average Current
        # Column 3: Duration
        # Column 4: Interval
        # Column 5: Average Power (uA) <- Key column
        # Column 6: Peak Current (mA)
        # Column 7: Notes

        current_scene = None

        for idx, row in df.iterrows():
            # Extract scene from column 0 (if present)
            if self.pd.notna(row.iloc[0]):
                current_scene = str(row.iloc[0]).strip()

            # Extract mode from column 1
            mode = None
            if self.pd.notna(row.iloc[1]):
                mode_str = str(row.iloc[1]).strip()
                # Combine scene and mode for better description
                if current_scene:
                    mode = f"{current_scene}_{mode_str}"
                else:
                    mode = mode_str

            # Extract current from column 5 (index 5) - "低功耗
            current_ua = 0.0
            if self.pd.notna(row.iloc[5]):
                current_ua = self._parse_current_value(str(row.iloc[5]))

            # Skip if no valid data
            if not mode or current_ua <= 0:
                continue

            # Calculate power
            voltage_mv = 3000  # Default 3V
            power_mw = (voltage_mv * current_ua) / 1000

            # Extract conditions from other columns
            conditions = []
            if self.pd.notna(row.iloc[3]):  # Duration
                conditions.append(f"duration: {row.iloc[3]}")
            if self.pd.notna(row.iloc[4]):  # Interval
                conditions.append(f"interval: {row.iloc[4]}")
            if self.pd.notna(row.iloc[7]):  # Notes
                conditions.append(f"note: {row.iloc[7]}")

            entry = PowerEntry(
                mode=self._normalize_mode(mode),
                peripheral="BLE",
                current_ua=current_ua,
                voltage_mv=voltage_mv,
                power_mw=power_mw,
                conditions="; ".join(conditions)
            )
            entries.append(entry)

        return entries

    def _parse_power_map_openpyxl(self, xlsx_path: Path) -> List[PowerEntry]:
        """Parse power map using openpyxl."""
        entries = []

        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Use first sheet
        sheet = wb.active

        # Iterate through rows starting from row 2 (skip headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            mode = self._extract_power_mode_from_row(row)
            current_ua = self._extract_current_ua_from_row(row)

            if mode and current_ua and current_ua > 0:
                voltage_mv = 3000  # Default 3V
                power_mw = (voltage_mv * current_ua) / 1000
                conditions = self._extract_conditions_from_row(row)

                entry = PowerEntry(
                    mode=mode,
                    peripheral="BLE",
                    current_ua=current_ua,
                    voltage_mv=voltage_mv,
                    power_mw=power_mw,
                    conditions=conditions
                )
                entries.append(entry)

        wb.close()
        return entries

    def _extract_power_mode(self, row) -> Optional[str]:
        """Extract power mode from pandas row."""
        # Check columns for mode information
        for col_name in ['场景', '模式', 'mode', 'Mode', 'Scene', 'scene']:
            if col_name in row.index:
                value = row[col_name]
                if value and str(value).strip():
                    return self._normalize_mode(str(value))

        return None

    def _extract_current_ua(self, row) -> float:
        """Extract current in microamps from pandas row."""
        # Check multiple columns for current values
        for col_name in ['平均功耗', '低功耗', '3V平均电流', '电流', 'Current',
                         'current_ua', 'Average Current']:
            if col_name in row.index:
                value = row[col_name]
                if value:
                    return self._parse_current_value(str(value))

        # Check all columns for patterns like "70uA", "3.3mA"
        for col_name in row.index:
            value = str(row[col_name]).strip()
            if any(unit in value.lower() for unit in ['ua', 'ma', 'a']):
                return self._parse_current_value(value)

        return 0.0

    def _extract_power_conditions(self, row) -> str:
        """Extract test conditions from pandas row."""
        conditions = []

        for col_name in ['备注', '持续时间', '间隔', 'Note', 'Conditions']:
            if col_name in row.index:
                value = row[col_name]
                if value and str(value).strip():
                    conditions.append(f"{col_name}: {value}")

        return "; ".join(conditions) if conditions else ""

    def _extract_power_mode_from_row(self, row) -> Optional[str]:
        """Extract power mode from openpyxl row."""
        # Check first few columns for mode
        for value in list(row)[:5]:
            if value and isinstance(value, str):
                mode = self._normalize_mode(value)
                if mode != value:  # If it was normalized
                    return mode
        return None

    def _extract_current_ua_from_row(self, row) -> Optional[float]:
        """Extract current from openpyxl row."""
        for value in row:
            if value:
                current = self._parse_current_value(str(value))
                if current > 0:
                    return current
        return None

    def _extract_conditions_from_row(self, row) -> str:
        """Extract conditions from openpyxl row."""
        conditions = []
        for value in row:
            if value and isinstance(value, str) and len(value) > 3:
                conditions.append(value)
        return "; ".join(conditions[:3]) if conditions else ""

    def _parse_current_value(self, value_str: str) -> float:
        """Parse current value string to microamps."""
        value_str = value_str.strip().upper()

        # Extract numeric part
        num_str = ''.join(c for c in value_str if c.isdigit() or c == '.')
        try:
            num = float(num_str)
        except ValueError:
            return 0.0

        # Convert to microamps
        if 'MA' in value_str:
            return num * 1000  # mA to uA
        elif 'UA' in value_str or 'A' in value_str:
            return num  # Already uA
        else:
            # Default to uA
            return num

    def _normalize_mode(self, mode_str: str) -> str:
        """Normalize mode string to standard format."""
        mode_str = mode_str.strip()

        mode_mapping = {
            'sleep': 'SLEEP',
            'sleep模式': 'SLEEP',
            'poweroff': 'DEEP_SLEEP',
            'poweroff模式': 'DEEP_SLEEP',
            '待机': 'SLEEP',
            '广播': 'ACTIVE_ADVERTISING',
            '连接': 'ACTIVE_CONNECTED',
            '发送': 'ACTIVE_TX',
            '接收': 'ACTIVE_RX',
        }

        mode_lower = mode_str.lower()
        for key, value in mode_mapping.items():
            if key in mode_lower:
                return value

        # Return uppercase if no match
        return mode_str.upper()

    # ========================================================================
    # BLE SRAM Allocation Parsing
    # ========================================================================

    def parse_ble_sram_allocation(self, xlsx_path: str) -> List[SRAMRegion]:
        """
        Parse B6x_BLE-Sram空间分配.xlsx to extract SRAM allocation for BLE.

        Expected Excel format:
        - Sheet: "BLE SRAM" or "SRAM分配"
        - Columns: 区域, 起始地址, 结束地址, 大小, 描述, 库类型

        Args:
            xlsx_path: Path to BLE SRAM allocation Excel file

        Returns:
            List of SRAMRegion objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"BLE SRAM allocation file not found: {xlsx_path}")
            return []

        regions = []

        try:
            if self._pandas_available:
                regions = self._parse_sram_allocation_pandas(xlsx_path)
            elif self._openpyxl_available:
                regions = self._parse_sram_allocation_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(regions)} SRAM regions from {xlsx_path.name}")
            return regions

        except Exception as e:
            logger.error(f"Failed to parse SRAM allocation {xlsx_path}: {e}")
            return []

    def _parse_sram_allocation_pandas(self, xlsx_path: Path) -> List[SRAMRegion]:
        """Parse SRAM allocation using pandas."""
        regions = []

        # Try different sheet names
        for sheet_name in ['BLE SRAM', 'SRAM分配', 'SRAM', 'Memory_Layout', 'Sheet1',
                           'SRAM空间分配', 'SRAM\u7a7a\u95f4\u5206\u914d']:
            try:
                df = self.pd.read_excel(xlsx_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        else:
            df = self.pd.read_excel(xlsx_path)

        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]

        # Find column mappings
        col_map = self._find_column_mapping(df.columns,
            ['region', 'area', '区域', 'start', 'begin', '起始',
             'end', 'stop', '结束', 'size', '大小', 'desc', '描述',
             'library', 'lib', '库'])

        for _, row in df.iterrows():
            try:
                # Extract region name
                region_name = ""
                for key in ['region', 'area', '区域']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val and val.lower() not in ['nan', '']:
                            region_name = val
                            break

                # Extract start address
                start_addr = ""
                for key in ['start', 'begin', '起始']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            start_addr = val
                            break

                # Extract end address
                end_addr = ""
                for key in ['end', 'stop', '结束']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            end_addr = val
                            break

                # Extract size
                size_kb = 0
                for key in ['size', '大小']:
                    if key in col_map:
                        val = row[col_map[key]]
                        if self.pd.notna(val):
                            size_kb = self._parse_size(val)
                            break

                # Extract description
                description = ""
                for key in ['desc', '描述']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            description = val
                            break

                # Extract library type
                lib_type = ""
                for key in ['library', 'lib', '库']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val and val.lower() not in ['nan', '']:
                            lib_type = val
                            break

                if region_name:
                    region = SRAMRegion(
                        region_name=region_name,
                        start_address=start_addr or "0x0",
                        end_address=end_addr or "0x0",
                        size_bytes=size_kb * 1024,
                        size_kb=size_kb,
                        description=description,
                        library_type=lib_type
                    )
                    regions.append(region)

            except Exception as e:
                logger.debug(f"Skipping SRAM row due to error: {e}")

        return regions

    def _parse_sram_allocation_openpyxl(self, xlsx_path: Path) -> List[SRAMRegion]:
        """Parse SRAM allocation using openpyxl."""
        regions = []

        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet_name in ['BLE SRAM', 'SRAM分配', 'SRAM',
                           'SRAM空间分配', 'SRAM\u7a7a\u95f4\u5206\u914d']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                break
        else:
            sheet = wb.active

        # Find header row
        header_row = None
        for row in sheet.iter_rows(max_row=5, values_only=True):
            if any(cell and ('region' in str(cell).lower() or '区域' in str(cell) or 'sram' in str(cell).lower()) for cell in row if cell):
                header_row = row
                break

        if not header_row:
            wb.close()
            return []

        # Build column mapping
        col_map = {}
        for i, cell in enumerate(header_row):
            if cell:
                col_lower = str(cell).lower()
                if 'region' in col_lower or 'area' in col_lower or '区域' in col_lower:
                    col_map['region'] = i
                elif 'start' in col_lower or 'begin' in col_lower or '起始' in col_lower:
                    col_map['start'] = i
                elif 'end' in col_lower or 'stop' in col_lower or '结束' in col_lower:
                    col_map['end'] = i
                elif 'size' in col_lower or '大小' in col_lower:
                    col_map['size'] = i
                elif 'desc' in col_lower or '描述' in col_lower:
                    col_map['desc'] = i
                elif 'lib' in col_lower or '库' in col_lower:
                    col_map['lib'] = i

        # Read data rows
        for row in sheet.iter_rows(min_row=header_row[0].row + 1, values_only=True):
            if not row or not row[0]:
                continue

            region_name = str(row[col_map.get('region', 0)]).strip() if col_map.get('region') else ""
            start_addr = str(row[col_map.get('start', 1)]).strip() if col_map.get('start') else ""
            end_addr = str(row[col_map.get('end', 2)]).strip() if col_map.get('end') else ""
            size_kb = self._parse_size(row[col_map.get('size', 3)]) if col_map.get('size') else 0
            description = str(row[col_map.get('desc', 4)]).strip() if col_map.get('desc') else ""
            lib_type = str(row[col_map.get('lib', 5)]).strip() if col_map.get('lib') else ""

            if region_name:
                region = SRAMRegion(
                    region_name=region_name,
                    start_address=start_addr,
                    end_address=end_addr,
                    size_bytes=size_kb * 1024,
                    size_kb=size_kb,
                    description=description,
                    library_type=lib_type
                )
                regions.append(region)

        wb.close()
        return regions

    # ========================================================================
    # BLE Compatibility Parsing
    # ========================================================================

    def parse_ble_compatibility(self, xlsx_path: str) -> List[CompatibilityEntry]:
        """
        Parse B6x BLE兼容性列表.xlsx to extract BLE feature compatibility.

        Expected Excel format:
        - Sheet: "Compatibility" or "兼容性"
        - Columns: 功能, 支持库, 最小SDK版本, 最大SDK版本, 备注, 约束

        Args:
            xlsx_path: Path to BLE compatibility Excel file

        Returns:
            List of CompatibilityEntry objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"BLE compatibility file not found: {xlsx_path}")
            return []

        entries = []

        try:
            if self._pandas_available:
                entries = self._parse_compatibility_pandas(xlsx_path)
            elif self._openpyxl_available:
                entries = self._parse_compatibility_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(entries)} compatibility entries from {xlsx_path.name}")
            return entries

        except Exception as e:
            logger.error(f"Failed to parse compatibility {xlsx_path}: {e}")
            return []

    def _parse_compatibility_pandas(self, xlsx_path: Path) -> List[CompatibilityEntry]:
        """Parse compatibility using pandas."""
        entries = []

        # Try different sheet names
        for sheet_name in ['Compatibility', '兼容性', 'Features', 'Sheet1',
                           'Table 1', 'Table1', 'Table']:
            try:
                df = self.pd.read_excel(xlsx_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        else:
            df = self.pd.read_excel(xlsx_path)

        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]

        # Find column mappings
        col_map = self._find_column_mapping(df.columns,
            ['feature', '功能', 'support', '支持', 'lib', '库',
             'min', '最小', 'max', '最大', 'version', '版本',
             'note', '备注', 'constraint', '约束'])

        for _, row in df.iterrows():
            try:
                # Extract feature name
                feature_name = ""
                for key in ['feature', '功能']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val and val.lower() not in ['nan', '']:
                            feature_name = val
                            break

                if not feature_name:
                    continue

                # Extract supported libraries
                supported_libs = []
                for key in ['support', '支持', 'lib', '库']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            # Parse comma-separated values
                            supported_libs = [lib.strip() for lib in val.replace(';', ',').split(',')]
                            break

                # Extract version info
                min_version = ""
                for key in ['min', '最小']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            min_version = val
                            break

                max_version = ""
                for key in ['max', '最大']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            max_version = val
                            break

                # Extract notes
                notes = ""
                for key in ['note', '备注']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            notes = val
                            break

                # Extract constraints
                constraints = []
                for key in ['constraint', '约束']:
                    if key in col_map:
                        val = str(row[col_map[key]]).strip() if self.pd.notna(row[col_map[key]]) else ""
                        if val:
                            constraints = [c.strip() for c in val.replace(';', ',').split(',')]
                            break

                entry = CompatibilityEntry(
                    feature_name=feature_name,
                    supported_libs=supported_libs,
                    min_sdk_version=min_version,
                    max_sdk_version=max_version,
                    notes=notes,
                    constraints=constraints
                )
                entries.append(entry)

            except Exception as e:
                logger.debug(f"Skipping compatibility row due to error: {e}")

        return entries

    def _parse_compatibility_openpyxl(self, xlsx_path: Path) -> List[CompatibilityEntry]:
        """Parse compatibility using openpyxl."""
        entries = []

        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet_name in ['Compatibility', '兼容性', 'Features',
                           'Table 1', 'Table1', 'Table']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                break
        else:
            sheet = wb.active

        # Find header row
        header_row = None
        for row in sheet.iter_rows(max_row=5, values_only=True):
            if any(cell and ('feature' in str(cell).lower() or '功能' in str(cell).lower() or 'compatibility' in str(cell).lower()) for cell in row if cell):
                header_row = row
                break

        if not header_row:
            wb.close()
            return []

        # Build column mapping
        col_map = {}
        for i, cell in enumerate(header_row):
            if cell:
                col_lower = str(cell).lower()
                if 'feature' in col_lower or '功能' in col_lower:
                    col_map['feature'] = i
                elif 'support' in col_lower or '支持' in col_lower or 'lib' in col_lower or '库' in col_lower:
                    col_map['support'] = i
                elif 'min' in col_lower or '最小' in col_lower:
                    col_map['min'] = i
                elif 'max' in col_lower or '最大' in col_lower:
                    col_map['max'] = i
                elif 'note' in col_lower or '备注' in col_lower:
                    col_map['note'] = i
                elif 'constraint' in col_lower or '约束' in col_lower:
                    col_map['constraint'] = i

        # Read data rows
        for row in sheet.iter_rows(min_row=header_row[0].row + 1, values_only=True):
            if not row or not row[0]:
                continue

            feature_name = str(row[col_map.get('feature', 0)]).strip() if col_map.get('feature') else ""
            if not feature_name:
                continue

            support_val = str(row[col_map.get('support', 1)]).strip() if col_map.get('support') else ""
            supported_libs = [lib.strip() for lib in support_val.replace(';', ',').split(',')] if support_val else []

            min_version = str(row[col_map.get('min', 2)]).strip() if col_map.get('min') else ""
            max_version = str(row[col_map.get('max', 3)]).strip() if col_map.get('max') else ""
            notes = str(row[col_map.get('note', 4)]).strip() if col_map.get('note') else ""

            constraint_val = str(row[col_map.get('constraint', 5)]).strip() if col_map.get('constraint') else ""
            constraints = [c.strip() for c in constraint_val.replace(';', ',').split(',')] if constraint_val else []

            entry = CompatibilityEntry(
                feature_name=feature_name,
                supported_libs=supported_libs,
                min_sdk_version=min_version,
                max_sdk_version=max_version,
                notes=notes,
                constraints=constraints
            )
            entries.append(entry)

        wb.close()
        return entries

    # ========================================================================
    # BLE Phone Compatibility Issues Parsing (Domain 3: BLE Stack)
    # ========================================================================

    def parse_ble_phone_compatibility_issues(self, xlsx_path: str) -> List[PhoneCompatibilityIssue]:
        """
        Parse B6x BLE兼容性列表.xlsx to extract specific phone model issues (坑点).

        Expected Excel format:
        - Sheet: "手机兼容性" or "Phone Compatibility" or "坑点"
        - Columns: 手机型号, 品牌, 系统版本, 问题描述, 根本原因, 解决方案, 临时方案, BLE版本, SDK版本, 严重程度

        Args:
            xlsx_path: Path to BLE compatibility Excel file

        Returns:
            List of PhoneCompatibilityIssue objects
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"BLE compatibility file not found: {xlsx_path}")
            return []

        issues = []

        try:
            if self._pandas_available:
                issues = self._parse_phone_issues_pandas(xlsx_path)
            elif self._openpyxl_available:
                issues = self._parse_phone_issues_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(issues)} phone compatibility issues from {xlsx_path.name}")
            return issues

        except Exception as e:
            logger.error(f"Failed to parse phone issues {xlsx_path}: {e}")
            return []

    def _parse_phone_issues_pandas(self, xlsx_path: Path) -> List[PhoneCompatibilityIssue]:
        """Parse phone compatibility issues using pandas."""
        issues = []

        # Try different sheet names
        for sheet_name in ['手机兼容性', 'Phone Compatibility', '坑点', 'Phone Issues', '兼容性',
                           'Table 1', 'Table1', 'Table']:
            try:
                df = self.pd.read_excel(xlsx_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        else:
            df = self.pd.read_excel(xlsx_path)

        df.columns = [str(col).strip() for col in df.columns]

        # Build column mapping
        col_map = self._find_column_mapping(df.columns,
            ['phone', 'model', '型号', 'brand', '品牌', 'android', 'ios', '系统',
             'issue', '问题', 'cause', '原因', 'solution', '解决', 'workaround', '临时',
             'ble', 'version', 'sdk', 'severity', '严重'])

        for _, row in df.iterrows():
            try:
                # Helper function to safely get cell value by column name
                def get_cell(col_key, default=""):
                    col_name = col_map.get(col_key)
                    if col_name and col_name in row.index:
                        val = row[col_name]
                        if self.pd.notna(val):
                            return str(val).strip()
                    return default

                phone_model = get_cell('phone')
                if not phone_model:
                    continue

                phone_brand = get_cell('brand')
                android_version = get_cell('android')
                issue_description = get_cell('issue')
                root_cause = get_cell('cause')
                solution = get_cell('solution')
                workaround = get_cell('workaround')
                affected_ble_version = get_cell('ble')
                sdk_version_required = get_cell('sdk')

                severity_val = get_cell('severity').lower()
                severity = "medium"
                if severity_val in ['critical', 'high', 'low', '严重', '高', '低']:
                    severity = severity_val

                issue = PhoneCompatibilityIssue(
                    phone_model=phone_model,
                    phone_brand=phone_brand,
                    android_version=android_version,
                    issue_type="兼容性问题",
                    issue_description=issue_description,
                    root_cause=root_cause,
                    workaround=workaround,
                    solution=solution,
                    affected_ble_version=affected_ble_version,
                    sdk_version_required=sdk_version_required,
                    severity=severity
                )
                issues.append(issue)

            except Exception as e:
                logger.debug(f"Skipping phone issue row: {e}")

        return issues

    def _parse_phone_issues_openpyxl(self, xlsx_path: Path) -> List[PhoneCompatibilityIssue]:
        """Parse phone compatibility issues using openpyxl (simplified)."""
        issues = []
        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            phone_model = str(row[0]).strip() if row[0] else ""
            if not phone_model:
                continue

            phone_brand = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            android_version = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            issue_description = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            root_cause = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            solution = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            workaround = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            issue = PhoneCompatibilityIssue(
                phone_model=phone_model,
                phone_brand=phone_brand,
                android_version=android_version,
                issue_type="兼容性问题",
                issue_description=issue_description,
                root_cause=root_cause,
                workaround=workaround,
                solution=solution,
                affected_ble_version="",
                sdk_version_required="",
                severity="medium"
            )
            issues.append(issue)

        wb.close()
        return issues

    # ========================================================================
    # BLE Memory Boundaries Parsing (Domain 1: Hardware & Registers)
    # ========================================================================

    def parse_ble_memory_boundaries(self, xlsx_path: str) -> List[MemoryBoundary]:
        """
        Parse B6x_BLE-Sram空间分配.xlsx to extract memory boundaries and limits.

        Args:
            xlsx_path: Path to BLE SRAM allocation Excel file

        Returns:
            List of MemoryBoundary objects with boundary/limit information
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.warning(f"BLE SRAM allocation file not found: {xlsx_path}")
            return []

        boundaries = []

        try:
            if self._pandas_available:
                boundaries = self._parse_memory_boundaries_pandas(xlsx_path)
            elif self._openpyxl_available:
                boundaries = self._parse_memory_boundaries_openpyxl(xlsx_path)
            else:
                logger.error("No Excel parsing library available")
                return []

            logger.info(f"Parsed {len(boundaries)} memory boundaries from {xlsx_path.name}")
            return boundaries

        except Exception as e:
            logger.error(f"Failed to parse memory boundaries {xlsx_path}: {e}")
            return []

    def _parse_memory_boundaries_pandas(self, xlsx_path: Path) -> List[MemoryBoundary]:
        """Parse memory boundaries using pandas."""
        boundaries = []

        for sheet_name in ['SRAM分配', 'Memory Layout', 'Sheet1',
                           'SRAM空间分配', 'SRAM\u7a7a\u95f4\u5206\u914d']:
            try:
                df = self.pd.read_excel(xlsx_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        else:
            df = self.pd.read_excel(xlsx_path)

        df.columns = [str(col).strip() for col in df.columns]

        for _, row in df.iterrows():
            try:
                region_name = str(row.iloc[0]).strip() if self.pd.notna(row.iloc[0]) else ""
                if not region_name:
                    continue

                start_addr = str(row.iloc[1]).strip() if len(row) > 1 and self.pd.notna(row.iloc[1]) else ""
                end_addr = str(row.iloc[2]).strip() if len(row) > 2 and self.pd.notna(row.iloc[2]) else ""
                size_kb = self._parse_size(row.iloc[3]) if len(row) > 3 and self.pd.notna(row.iloc[3]) else 0
                boundary_type = str(row.iloc[4]).strip() if len(row) > 4 and self.pd.notna(row.iloc[4]) else ""
                lib_type = str(row.iloc[5]).strip() if len(row) > 5 and self.pd.notna(row.iloc[5]) else ""
                description = str(row.iloc[6]).strip() if len(row) > 6 and self.pd.notna(row.iloc[6]) else ""

                boundary = MemoryBoundary(
                    region_name=region_name,
                    start_address=start_addr,
                    end_address=end_addr,
                    size_bytes=size_kb * 1024,
                    size_kb=size_kb,
                    boundary_type=boundary_type,
                    library_type=lib_type,
                    description=description,
                    access_restrictions=[],
                    overlap_warnings=[],
                    alignment_requirement=4
                )
                boundaries.append(boundary)

            except Exception as e:
                logger.debug(f"Skipping memory boundary row: {e}")

        return boundaries

    def _parse_memory_boundaries_openpyxl(self, xlsx_path: Path) -> List[MemoryBoundary]:
        """Parse memory boundaries using openpyxl (simplified)."""
        boundaries = []
        wb = self.openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            region_name = str(row[0]).strip() if row[0] else ""
            if not region_name:
                continue

            start_addr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            end_addr = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            size_kb = self._parse_size(row[3]) if len(row) > 3 and row[3] else 0
            boundary_type = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            lib_type = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            description = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            boundary = MemoryBoundary(
                region_name=region_name,
                start_address=start_addr,
                end_address=end_addr,
                size_bytes=size_kb * 1024,
                size_kb=size_kb,
                boundary_type=boundary_type,
                library_type=lib_type,
                description=description,
                access_restrictions=[],
                overlap_warnings=[],
                alignment_requirement=4
            )
            boundaries.append(boundary)

        wb.close()
        return boundaries



def parse_io_map_file(xlsx_path: str) -> List[PinMuxEntry]:
    """Convenience function to parse IO map file."""
    parser = ExcelHardwareParser()
    return parser.parse_io_map(xlsx_path)


def parse_flash_map_file(xlsx_path: str) -> List[MemoryRegion]:
    """Convenience function to parse flash map file."""
    parser = ExcelHardwareParser()
    return parser.parse_flash_map(xlsx_path)


def parse_power_map_file(xlsx_path: str) -> List[PowerEntry]:
    """Convenience function to parse power consumption file."""
    parser = ExcelHardwareParser()
    return parser.parse_power_map(xlsx_path)


def parse_ble_sram_allocation_file(xlsx_path: str) -> List[SRAMRegion]:
    """Convenience function to parse BLE SRAM allocation file."""
    parser = ExcelHardwareParser()
    return parser.parse_ble_sram_allocation(xlsx_path)


def parse_ble_compatibility_file(xlsx_path: str) -> List[CompatibilityEntry]:
    """Convenience function to parse BLE compatibility file."""
    parser = ExcelHardwareParser()
    return parser.parse_ble_compatibility(xlsx_path)


def parse_ble_phone_compatibility_issues_file(xlsx_path: str) -> List[PhoneCompatibilityIssue]:
    """Convenience function to parse BLE phone compatibility issues (坑点)."""
    parser = ExcelHardwareParser()
    return parser.parse_ble_phone_compatibility_issues(xlsx_path)


def parse_ble_memory_boundaries_file(xlsx_path: str) -> List[MemoryBoundary]:
    """Convenience function to parse BLE memory boundaries."""
    parser = ExcelHardwareParser()
    return parser.parse_ble_memory_boundaries(xlsx_path)


if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format='%(levelname)s: %(message)s'
    )

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <excel_file>")
        sys.exit(1)

    parser = ExcelHardwareParser()
    file_path = sys.argv[1]

    if 'io' in file_path.lower() or 'pin' in file_path.lower():
        entries = parser.parse_io_map(file_path)
        print(f"\nFound {len(entries)} pin mux entries:")
        for entry in entries[:10]:
            print(f"  {entry.pin_number}: {entry.alternate_function} ({entry.peripheral})")

    elif 'flash' in file_path.lower() or 'memory' in file_path.lower():
        regions = parser.parse_flash_map(file_path)
        print(f"\nFound {len(regions)} memory regions:")
        for region in regions[:10]:
            print(f"  {region.region_name}: {region.base_address} ({region.size_kb}KB)")
